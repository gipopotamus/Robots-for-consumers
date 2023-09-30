import os
import tempfile
from openpyxl import load_workbook
from django.test import TestCase
from django.urls import reverse
from django.http import HttpResponse
from .models import Robot
from .services import generate_excel_report
from django.utils import timezone


class GenerateExcelReportTestCase(TestCase):
    def setUp(self):
        # Создаем несколько объектов Robot для тестирования
        current_date = timezone.now()
        self.robot1 = Robot.objects.create(
            model='R2',
            version='D2',
            serial='R2D2',
            created=current_date
        )
        self.robot2 = Robot.objects.create(
            model='R2',
            version='A1',
            serial='R2A1',
            created=current_date
        )

    def test_generate_excel_report_success(self):
        # Вызываем функцию generate_excel_report
        response = generate_excel_report(HttpResponse())

        # Проверяем, что ответ имеет статус 200 OK
        self.assertEqual(response.status_code, 200)

        # Создаем временный файл для сохранения данных из ответа
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file.write(response.content)

        # Загружаем сгенерированный Excel-файл из временного файла
        generated_workbook = load_workbook(temp_file.name)

        # Проверяем, что в файле есть ожидаемые листы
        self.assertIn('R2 - D2', generated_workbook.sheetnames)
        self.assertIn('R2 - A1', generated_workbook.sheetnames)

        # Затем удаляем временный файл
        os.remove(temp_file.name)
